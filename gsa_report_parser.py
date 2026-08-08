#!/usr/bin/env python3
"""Parse GSA (Global Solar Atlas) XLSX report files.

Supports both:
1. Pure meteorological report (DNI profile only)
2. Full PV report (with PV system config, PVOUT hourly profile)

Usage:
    python gsa_report_parser.py <path_to_xlsx> [--format json|table|hourly]
    
Output: structured data including hourly DNI profiles and PVOUT (if available).

Example:
    python gsa_report_parser.py ~/Downloads/GSA_Report_Linping_District.xlsx
"""

import sys
import json
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_overview(ws):
    """Parse Overview sheet."""
    data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[0] and row[1] is not None:
            key = str(row[0]).strip()
            val = str(row[1]).strip() if row[1] else ""
            data[key] = val
    return data


def parse_site_info(ws):
    """Parse Site_info sheet."""
    data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[0] and row[1] is not None:
            key = str(row[0]).strip()
            val = str(row[1]).strip() if row[1] else ""
            data[key] = val
    return data


def parse_pv_config(ws):
    """Parse PV_config sheet (present in full PV reports)."""
    data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[0] and row[1] is not None:
            key = str(row[0]).strip()
            val = row[1]
            unit = str(row[2]).strip() if row[2] else ""
            data[key] = {"value": val, "unit": unit}
    return data


def parse_map_data(ws):
    """Parse Map_data sheet."""
    data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[0] and row[2] is not None:
            name = str(row[0]).strip()
            acronym = str(row[1]).strip() if row[1] else ""
            value = row[2]
            unit = str(row[3]).strip() if row[3] else ""
            data[name] = {
                "acronym": acronym,
                "value": value,
                "unit": unit
            }
    return data


def parse_monthly_averages(ws):
    """Parse Monthly_averages sheet.
    
    Supports two formats:
    1. Simple: only DNI column
    2. Full: PVOUT_specific + PVOUT_total + DNI
    """
    data = {}
    headers = []
    found_header = False
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not found_header:
            # Detect header row with column names
            # Row[0] can be None or empty string ''
            if (row[0] is None or str(row[0]).strip() == '') and any(h in str(row[1] or "") for h in ["PVOUT", "DNI"]):
                headers = [str(row[i]).strip() if row[i] else "" for i in range(len(row))]
                found_header = True
            continue
        
        if row[0]:
            month = str(row[0]).strip()
            data[month] = {}
            for i, h in enumerate(headers[1:], start=1):
                if h and row[i] is not None:
                    val = row[i]
                    # Handle string numbers with commas (e.g. "69,186.7")
                    if isinstance(val, str):
                        val = val.replace(",", "").strip()
                        try:
                            val = float(val)
                        except ValueError:
                            pass
                    data[month][h] = val
    
    return data


def parse_hourly_profiles(ws):
    """Parse Hourly_profiles sheet - supports both pure DNI and full PV report.
    
    Full PV report has TWO sections:
    1. Total photovoltaic power output [Wh] (PVOUT_total)
    2. Direct normal irradiation [Wh/m²] (DNI)
    
    Returns dict with:
        - has_pvout: bool
        - pvout_total: dict (if present) - {month: {hour: value}}
        - dni: dict - {month: {hour: value}}
        - pvout_sum: dict (if present)
        - dni_sum: dict
    """
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    
    result = {
        "hours": [],
        "months": months,
        "has_pvout": False,
        "pvout_total": {},
        "dni": {},
        "pvout_sum": {},
        "dni_sum": {},
    }
    
    current_section = None  # "pvout" or "dni"
    found_months = False
    hour_labels = []
    pvout_data = {m: {} for m in months}
    dni_data = {m: {} for m in months}
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cells = list(row)
        
        # Detect section headers
        header_str = " ".join(str(c or "") for c in cells)
        if "photovoltaic power output" in header_str.lower():
            current_section = "pvout"
            result["has_pvout"] = True
            found_months = False
            continue
        elif "direct normal irradiation" in header_str.lower():
            current_section = "dni"
            found_months = False
            continue
        
        # Detect month header row
        if not found_months:
            if len(cells) >= 12 and cells[1] in months:
                found_months = True
                continue
            continue
        
        if not row[0]:
            continue
        
        label = str(row[0]).strip()
        
        if label.lower() == "sum":
            # Sum row
            for i, m in enumerate(months):
                if row[i+1] is not None:
                    if current_section == "pvout":
                        result["pvout_sum"][m] = row[i+1]
                    elif current_section == "dni":
                        result["dni_sum"][m] = row[i+1]
            continue
        
        # Hour row - only collect labels once (from first section)
        if current_section == "pvout":
            hour_labels.append(label)
            for i, m in enumerate(months):
                if row[i+1] is not None:
                    pvout_data[m][label] = row[i+1]
        elif current_section == "dni":
            for i, m in enumerate(months):
                if row[i+1] is not None:
                    dni_data[m][label] = row[i+1]
    
    result["hours"] = hour_labels
    result["pvout_total"] = pvout_data
    result["dni"] = dni_data
    
    return result


def parse_gsa_report(filepath):
    """Parse a GSA XLSX report file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    result = {
        "filename": Path(filepath).name,
        "sheets": {}
    }
    
    parsers = {
        "Overview": parse_overview,
        "Site_info": parse_site_info,
        "PV_config": parse_pv_config,
        "Map_data": parse_map_data,
        "Monthly_averages": parse_monthly_averages,
        "Hourly_profiles": parse_hourly_profiles,
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parser = parsers.get(sheet_name)
        if parser:
            try:
                result["sheets"][sheet_name] = parser(ws)
            except Exception as e:
                result["sheets"][sheet_name] = {"error": str(e)}
        else:
            # Generic read for unknown sheets
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            result["sheets"][sheet_name] = rows
    
    wb.close()
    return result


def format_hourly_table(hourly_data, month=None, show_dni=True):
    """Format hourly profiles as a readable table."""
    lines = []
    
    if hourly_data["has_pvout"]:
        # Full PV report - show both PVOUT and DNI
        lines.append("### Hourly PVOUT Profile [Wh]")
        lines.append("")
        months_to_show = [month] if month else hourly_data["months"]
        
        for m in months_to_show:
            lines.append(f"#### {m} - Total photovoltaic power output [Wh]")
            lines.append("")
            lines.append("| Hour | PVOUT |")
            lines.append("|------|-------|")
            for hour in hourly_data["hours"]:
                val = hourly_data["pvout_total"].get(m, {}).get(hour, 0)
                lines.append(f"| {hour:>5} | {val:>7.1f} |")
            lines.append("")
        
        if show_dni:
            lines.append("### Hourly DNI Profile [Wh/m²]")
            lines.append("")
            for m in months_to_show:
                lines.append(f"#### {m} - Direct Normal Irradiation [Wh/m²]")
                lines.append("")
                lines.append("| Hour | DNI |")
                lines.append("|------|-----|")
                for hour in hourly_data["hours"]:
                    val = hourly_data["dni"].get(m, {}).get(hour, 0)
                    lines.append(f"| {hour:>5} | {val:>6.1f} |")
                lines.append("")
    else:
        # Pure meteorological report - only DNI
        lines.append("### Hourly DNI Profile [Wh/m²]")
        lines.append("")
        months_to_show = [month] if month else hourly_data["months"]
        
        header = "| Hour | " + " | ".join(months_to_show) + " |"
        sep = "|------|" + "|".join(["------" for _ in months_to_show]) + "|"
        lines.append(header)
        lines.append(sep)
        
        for hour in hourly_data["hours"]:
            vals = []
            for m in months_to_show:
                v = hourly_data["dni"].get(m, {}).get(hour, 0)
                vals.append(f"{v:>6.1f}")
            lines.append(f"| {hour:>5} | " + " | ".join(vals) + " |")
    
    return "\n".join(lines)


def format_monthly_table(monthly_data):
    """Format monthly averages."""
    lines = []
    lines.append("### Monthly Averages")
    lines.append("")
    
    # Detect available columns
    sample = next(iter(monthly_data.values()), {})
    columns = list(sample.keys()) if isinstance(sample, dict) else []
    
    if not columns:
        # Simple format (single DNI column)
        lines.append("| Month | DNI (kWh/m²) |")
        lines.append("|-------|-------------|")
        for m in ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Yearly"]:
            if m in monthly_data:
                lines.append(f"| {m:>5} | {monthly_data[m]} |")
        return "\n".join(lines)
    
    # Full format with multiple columns
    header = "| Month | " + " | ".join(columns) + " |"
    sep = "|-------|" + "|".join(["-------" for _ in columns]) + "|"
    lines.append(header)
    lines.append(sep)
    
    for m in ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
             "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Yearly"]:
        if m in monthly_data:
            vals = []
            for col in columns:
                v = monthly_data[m].get(col, "")
                vals.append(f"{v:>10}" if v != "" else "          ")
            lines.append(f"| {m:>5} | " + " | ".join(vals) + " |")
    
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Parse GSA XLSX report")
    parser.add_argument("filepath", help="Path to GSA XLSX file")
    parser.add_argument("--format", choices=["json", "table", "hourly"], default="json",
                       help="Output format")
    parser.add_argument("--month", help="Show specific month only (e.g. Jul)")
    parser.add_argument("--no-dni", action="store_true", help="Skip DNI in hourly output")
    
    args = parser.parse_args()
    
    if not Path(args.filepath).exists():
        print(f"ERROR: File not found: {args.filepath}", file=sys.stderr)
        sys.exit(1)
    
    result = parse_gsa_report(args.filepath)
    
    if args.format == "json":
        print(json.dumps(result, indent=2, ensure_ascii=False))
    elif args.format == "hourly":
        hourly = result["sheets"].get("Hourly_profiles", {})
        print(format_hourly_table(hourly, args.month, show_dni=not args.no_dni))
    elif args.format == "table":
        # Pretty print all data
        map_data = result["sheets"].get("Map_data", {})
        print("## Map Data (Annual)")
        print()
        for name, info in map_data.items():
            print(f"- **{name}** ({info['acronym']}): {info['value']} {info['unit']}")
        
        pv_config = result["sheets"].get("PV_config", {})
        if pv_config:
            print()
            print("## PV System Configuration")
            print()
            for key, info in pv_config.items():
                print(f"- **{key}**: {info['value']} {info['unit']}")
        
        monthly = result["sheets"].get("Monthly_averages", {})
        if monthly:
            print()
            print(format_monthly_table(monthly))
        
        hourly = result["sheets"].get("Hourly_profiles", {})
        if hourly:
            print()
            print(format_hourly_table(hourly, args.month, show_dni=not args.no_dni))


if __name__ == "__main__":
    main()
